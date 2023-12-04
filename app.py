from flask import Flask, send_file, send_from_directory, render_template, request, jsonify
import pandas as pd
import openpyxl
from fuzzywuzzy import process
import os
import boto3
from io import BytesIO

app = Flask(__name__)

aws_access_key_id = os.getenv('BUCKETEER_AWS_ACCESS_KEY_ID')
aws_secret_access_key = os.getenv('BUCKETEER_AWS_SECRET_ACCESS_KEY')
bucket_name = os.getenv('BUCKETEER_BUCKET_NAME')


@app.route('/download_excel')
def download_excel():
    s3_client = boto3.client('s3', aws_access_key_id=aws_access_key_id, aws_secret_access_key=aws_secret_access_key)
    file_key = 'assets/SEMI_data.xlsx'

    try:
        response = s3_client.get_object(Bucket=bucket_name, Key=file_key)
        excel_data = response['Body'].read()

        return send_file(
            BytesIO(excel_data),
            as_attachment=True,
            download_name='SEMI_data.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return str(e)

@app.route('/')
def display_excel():
    s3_client = boto3.client('s3', aws_access_key_id=aws_access_key_id, aws_secret_access_key=aws_secret_access_key)
    file_key = 'assets/SEMI_data.xlsx'

    response = s3_client.get_object(Bucket=bucket_name, Key=file_key)
    excel_data = response['Body'].read()
    df = pd.read_excel(BytesIO(excel_data))

    table_html = df.to_html(classes='excel-table', border=0)
    return render_template('index.html', table_html=table_html)



@app.route('/update_excel', methods=['POST'])
def update_excel():

    s3_client = boto3.client(
        's3',
        aws_access_key_id=aws_access_key_id,
        aws_secret_access_key=aws_secret_access_key,
        region_name='us-east-1'
    )
    file_key = 'assets/SEMI_data.xlsx'

    response = s3_client.get_object(Bucket=bucket_name, Key=file_key)
    excel_data = response['Body'].read()
    workbook = openpyxl.load_workbook(BytesIO(excel_data))
    sheet = workbook.active


    metrics_to_columns = {
    "Total water consumed": "B",
    "Municipal water usage": "C",
    "Surface water usage": "D",
    "Groundwater usage": "E",
    "Water restored": "F",
    "Water reclaimed/reused": "G",
    "Water discharged": "H",
    "Quality of water discharged": "I",
    "Non-hazardous waste generated": "J",
    "Hazardous waste generated": "K",
    "Waste recycled (onsite or offsite)": "L",
    "Waste sent to landfill (hazardous and non-hazardous)": "M",
    "Waste incinerated (also referred to as 'energy recovery')": "N"
    }
    
    data = request.json.get('data')
    formatted_data = jsonifyData(data)
    



    company_name = formatted_data.get("Company Name")

    companies = [sheet.cell(row=row, column=1).value for row in range(1, sheet.max_row + 1)]
    
    # Use fuzzy matching to find the closest match to the given company name
    closest_company_name = find_best_match(company_name, companies)

    print(company_name)
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == closest_company_name:
            company_row = row
            break

    for key, value in formatted_data.items():
        if key in metrics_to_columns:
          col = metrics_to_columns[key]
          cell = f"{col}{company_row}"
          sheet[cell] = value

   



    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    s3_client.put_object(Bucket=bucket_name, Key=file_key, Body=output)

    # Convert the updated Excel file to a DataFrame and then to HTML
    output.seek(0)
    df = pd.read_excel(BytesIO(output.getvalue()))
    updated_table = df.to_html(classes='excel-table', border=0)

    # Return the updated table
    return jsonify(updatedTable=updated_table)

def find_best_match(name, choices):
    best_match = process.extractOne(name, choices)
    return best_match[0]  # returns the best match

def jsonifyData(data):
    lines = data.split("\n")
    result = {}
    company = lines[0].strip()
    
    result["Company Name"] = company  

    for line in lines[2:]:
        if ": " in line: 
            key, value = line.split(": ", 1)
            result[key] = value

    return result


if __name__ == '__main__':
    app.run(debug=True)
